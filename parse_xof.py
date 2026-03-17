import openpyxl
import re
import json
from collections import Counter, defaultdict
from datetime import datetime

# XOF/USD - XOF is pegged to EUR (655.957 XOF = 1 EUR)
# Using approximate rate: 1 USD ~ 605 XOF (as of early 2025)
XOF_PER_USD = 605.0

HEADERS = ['InputDate','PostedDate','AccountNumber','Account','Bank','AccountType','ValueDate','TransRef','Currency','Amount','BAIType','Comment']

# Country mapping by bank name and account prefix
BANK_COUNTRY = {
    'Bank Atlantique IC': 'Ivory Coast',
    'BOA Ivory Coast': 'Ivory Coast',
    'SOCGEN Ivory Coast': 'Ivory Coast',
    'SIB (IVORIAN BANK)': 'Ivory Coast',
    'CITI IVORY COAST': 'Ivory Coast',
    'STANBIC IVORY COAST': 'Ivory Coast',
    'MTN Mobile FS CI': 'Ivory Coast',
    'BICIS Senegal': 'Senegal',
    'Bank Atlantique SN': 'Senegal',
    'C.B.A.O. Senegal': 'Senegal',
    'Bank Africa Senegal': 'Senegal',
    'SOCGEN SENEGAL BNK': 'Senegal',
    'Citi Senegal Bank': 'Senegal',
    'AFG BK': 'Ivory Coast',       # AFG operates in CI; account prefix CI confirms
    'BANQUE POPULAIRE': 'Ivory Coast',  # Banque Populaire de Cote d'Ivoire
}

def get_country(bank_name, account_number):
    """Determine country from bank name, fallback to account number prefix."""
    if bank_name in BANK_COUNTRY:
        return BANK_COUNTRY[bank_name]
    acct = str(account_number or '')
    if acct.startswith('CI'):
        return 'Ivory Coast'
    elif acct.startswith('SN'):
        return 'Senegal'
    return 'Unknown'

def parse_row(row):
    d = {}
    for i, h in enumerate(HEADERS):
        if i < len(row):
            d[h] = row[i]
        else:
            d[h] = None
    return d

def _strip_company_suffix(text):
    """Remove trailing company name suffixes that aren't useful as beneficiary."""
    # Strip common company suffixes at end
    text = re.sub(r'(?:GRANDS MOULINS D[E\']? (?:ABIDJAN|DAKAR|DKR)|GRANDS MOULINS DE DAKAR|'
                  r'LES GRANDS MOULINS D[E\']? (?:ABIDJAN|DAKAR)|GMA|GMD)\s*$', '', text, flags=re.IGNORECASE).strip()
    return text

def _extract_beneficiary(fr, py, comment):
    """Extract counterparty/beneficiary from comment fields."""
    text = py if py else fr
    text_lower = text.lower()
    comment_lower = comment.lower()

    # VE EFFECTUE PAR: <person> — cash deposit by named person
    m = re.search(r've effectue par:\s*(.+)', text, re.IGNORECASE)
    if m:
        return _strip_company_suffix(m.group(1).strip())[:60]

    # VERSEMENT <name>Motif: or VERSEMENT <name>REMETTANT: — depositor name
    m = re.search(r'versement\s+(.+?)(?:motif:|remettant:|$)', text, re.IGNORECASE)
    if m:
        name = m.group(1).strip()
        # Only if it looks like a name (not a keyword like "especes", "deplace")
        if name and not any(x in name.lower() for x in ['espece', 'deplace', 'cfi', 'sur ', 'par ']):
            return _strip_company_suffix(name)[:60]

    # VIB FAV. <beneficiary> or VIB FAV <beneficiary>
    m = re.search(r'vib fav\.?\s+(.+?)(?:cion|$)', text, re.IGNORECASE)
    if m:
        return _strip_company_suffix(m.group(1).strip())[:60]

    # VIRMENT/VIREMENT RTGS FAVEUR <beneficiary>
    m = re.search(r'vir(?:e?ment)?\s+rtgs\s+faveur\s+(.+)', text, re.IGNORECASE)
    if m:
        return _strip_company_suffix(m.group(1).strip())[:60]

    # VIR FAVEUR / VIR BOAWEB FAV — beneficiary after keyword
    m = re.search(r'(?:vir(?:ement)?\s+(?:boaweb\s+)?faveur?|vir faveur)\s*(?:beneficiaire\s+)?(.+)', text, re.IGNORECASE)
    if m:
        return _strip_company_suffix(m.group(1).strip())[:60]

    # TRF EMIS <beneficiary> — outgoing transfer
    m = re.search(r'trf\s+emis\s+(.+?)(?:transfert|motif:|$)', text, re.IGNORECASE)
    if m:
        name = m.group(1).strip()
        if name and name.lower() not in ['eurafrique', '']:
            return _strip_company_suffix(name)[:60]

    # Transfert mis faveur: <beneficiary>
    m = re.search(r'(?:transfert|trf)\s+.{0,10}faveur[:\s]+(.+?)(?:motif:|montant:|$)', text, re.IGNORECASE)
    if m:
        return _strip_company_suffix(m.group(1).strip())[:60]

    # VT ETRANGER FAVEUR — foreign transfer, beneficiary is in FR:
    if 'vt etranger' in text_lower:
        fr_clean = _strip_company_suffix(fr)
        if fr_clean and fr_clean.lower() not in ['', 'nonref']:
            return fr_clean[:60]

    # VIR MASSE / VIR MULTI — mass payment, beneficiary in FR:
    if any(x in text_lower for x in ['vir masse', 'vir multi', 'virt multiple']):
        fr_clean = _strip_company_suffix(fr)
        if fr_clean and fr_clean.lower() not in ['', 'nonref']:
            return fr_clean[:60]

    # 05VIR.RECU: <sender> — incoming wire
    m = re.search(r'(?:05)?vir\.recu:\s*(.+)', text, re.IGNORECASE)
    if m:
        return _strip_company_suffix(m.group(1).strip())[:60]

    # EFFET AU <date> TIRE: <drawer>
    m = re.search(r'tire:\s*(.+)', text, re.IGNORECASE)
    if m:
        return _strip_company_suffix(m.group(1).strip())[:60]

    # /PT/FT/PY<description> — BICIS fee format, sometimes has person name
    m = re.search(r'/pt/ft/py(.+)', text, re.IGNORECASE)
    if m:
        name = m.group(1).strip()
        # Only treat as beneficiary if it looks like a name (not a fee/rejection description)
        if not any(x in name.lower() for x in ['frais', 'commission', 'timbre', 'telecompens',
                                                  'abonnement', 'facturation', 'bicis', 'agios',
                                                  'imp chq', 'miscellaneous', 'absence de',
                                                  'insuffisance', 'donnees faciales']):
            return _strip_company_suffix(name)[:60]

    # ESPECES VERSEES PAR <person> — name might be after PAR in text, or in FR: field
    m = re.search(r'especes?\s+vers[eé]e?s?\s+par\s+(.+)', text, re.IGNORECASE)
    if m:
        name = m.group(1).strip()
        if name and name.lower() not in ['par', '']:
            return _strip_company_suffix(name)[:60]
    # If PY: ends with "PAR" (name is cut off), use FR: field as the depositor name
    if 'especes versees par' in text_lower and text_lower.rstrip().endswith('par'):
        fr_clean = _strip_company_suffix(fr)
        if fr_clean and fr_clean.lower() not in ['', 'nonref']:
            return fr_clean[:60]

    # VERSEMENT ESPECES DE <person>
    m = re.search(r'versement\s+especes?\s+de\s+(.+)', text, re.IGNORECASE)
    if m:
        return _strip_company_suffix(m.group(1).strip())[:60]

    # REMETTANT: <person> — depositor in structured comments
    m = re.search(r'remettant:\s*(.+)', comment, re.IGNORECASE)
    if m:
        return _strip_company_suffix(m.group(1).strip())[:60]

    return ''

def classify_comment(comment):
    if not comment:
        return 'Unclassified', '', ''

    comment = str(comment)
    fr = ''
    trid = ''
    py = ''

    def _clean_py(raw):
        """Clean PY: field — strip company suffixes, Motif:, Remettant:, Montant: suffixes."""
        raw = re.split(r'(?:Motif:|REMETTANT:|Montant:|Au cours de:)', raw, flags=re.IGNORECASE)[0]
        return _strip_company_suffix(raw).strip()

    # --- Parse structured tags ---
    if 'FR:' in comment:
        parts = comment.split('FR:')
        rest = parts[1] if len(parts) > 1 else ''
        for tag in ['ENDT:', 'TRID:', 'PY:']:
            if tag in rest:
                idx = rest.index(tag)
                if not fr:
                    fr = rest[:idx].strip()
                rest = rest[idx:]
        if not fr:
            fr = rest.strip()
        if 'TRID:' in comment:
            trid = comment.split('TRID:')[1].split('PY:')[0].strip()
        if 'PY:' in comment:
            py = _clean_py(comment.split('PY:')[1])
    elif 'PY:' in comment:
        py = _clean_py(comment.split('PY:')[1])
        fr = comment.split('PY:')[0].strip()
    elif 'ENDT:' in comment or 'TRID:' in comment:
        # Has tags but no FR:/PY: — extract what we can
        fr = re.split(r'(?:ENDT:|TRID:)', comment)[0].strip()
        if 'PY:' in comment:
            py = _clean_py(comment.split('PY:')[1])
    else:
        fr = comment.strip()

    description = py if py else fr
    # Extract beneficiary
    bnf = _extract_beneficiary(fr, py, comment)
    # Clean up bad beneficiaries
    if bnf:
        bnf_lower = bnf.lower()
        # Filter out non-name artifacts
        if any(x in bnf_lower for x in ['miscellaneous swf', 'imp chq', 'absence de',
                                          'insuffisance', 'donnees faciales']):
            bnf = ''

    # --- Classify using BOTH fr and py fields ---
    fr_lower = fr.lower()
    py_lower = py.lower()
    # Combined text for broader matching
    combined = (fr_lower + ' ' + py_lower).strip()

    # --- CATEGORY RULES (ordered by specificity) ---

    # Tax / Stamp Duty — check before fees since "timbre" is tax, not a fee
    if any(x in combined for x in ['tax amount', 'tva ', 'impot', 'taxe ', 'droit timbre', 'timbre/vers',
                                     'timbre fiscal', 'timbre etat', 'timbre', 'vat']):
        cat = 'Tax'

    # Payroll / Salary
    elif any(x in combined for x in ['salary', 'salaire', 'paie ', 'paye ', 'cnps', 'paiement cnps']):
        cat = 'Payroll / Salary'

    # Foreign / Cross-border Transfers
    elif any(x in combined for x in ['vt etranger', 'trf emis eurafrique', 'frais trf eurafrique',
                                       'outward', 'inward', 'transfert international',
                                       'vor swft fav', 'vt interb']):
        cat = 'Cross-border'

    # Interbank Wire Transfers (VIB = Virement Interbancaire)
    elif any(x in combined for x in ['vib fav', 'cion / vi', 'virement interbancaire',
                                       'cion / virement']):
        cat = 'Wire Transfer'

    # Wire Transfer (general)
    elif any(x in combined for x in ['virement', 'vir.compense', 'vir int', 'vir emis', 'vir recu',
                                       'virt ', 'vir faveur', 'vir boaweb', 'vir masse',
                                       'vir multi', 'transfer', '05vir.recu:', '13vir boaweb',
                                       'virment rtgs', 'trf emis', 'vir ordre', 'virint ',
                                       'virmt ', 'virmt de ', 'vrt recu', 'vrt fav',
                                       'trasfert', '13vrt recu', '13virint']):
        cat = 'Wire Transfer'

    # Check / Cheque
    elif any(x in combined for x in ['cheque', 'chq ', 'chq.', 'chq/', 'remise chq', 'rem chq',
                                       'remise cheque', 'votre remose', 'chq comp', 'chq imp',
                                       'v/rem chq', 'remose chq', 'rejet sequence',
                                       'rej ch ', 'rem ch ']):
        cat = 'Check / Cheque'

    # Bills / Effets (commercial paper)
    elif any(x in combined for x in ['effet au', 'effet ', 'traite', 'bill of exchange',
                                       'encaissement effet', 'paiem.effets',
                                       'enct eff', 'regul enct', 'imp eff', 'impay ']):
        cat = 'Bills / Effets'

    # Cash Deposits/Withdrawals
    elif any(x in combined for x in ['espece', 'especes', 'retrait esp', 'versement esp',
                                       've effectue par', 'vers.esp', 'depot especes',
                                       'versement deplace', 'versement cfi', '/versement cfi',
                                       'borne ', 'cash', 'caisse', 'especes versees',
                                       'sort de caisse', 'retrait chq depl',
                                       'versement par to', 'versement sur',
                                       'ver dplace', 'ver deplace', 'retdab ']):
        cat = 'Cash'

    # Interest / Agios
    elif any(x in combined for x in ['agios', 'interest', 'interet']):
        cat = 'Interest / Agios'

    # Bank Fees / Commissions (broad — catches /PT/FT, telecompense fees, etc.)
    elif any(x in combined for x in ['commission', 'frais ', 'frais/', 'fees', 'charges',
                                       'com encais', 'com/', '+commissions',
                                       'frais sur', 'fraos', 'frais telecompense',
                                       'frais de t', 'frais de f', 'frais virement',
                                       'frais/virt', 'frais annuel', 'frais imp',
                                       'facturation', 'abonnement', 'pack anet',
                                       'pack ibe', 'sibnet', 'balance requirement',
                                       'debit chgs', 'number of debit', 'number of credit',
                                       '/pt/ft', 'minimum balance', 'taf sur loc',
                                       'bp online', 'pack anet', 't.p.s',
                                       'fraisfinex', 'cions de garde', 'loyer  ',
                                       'tob loyer']):
        cat = 'Fees / Commissions'

    # SWIFT / Telex
    elif any(x in combined for x in ['telex', 'swift', 'infoswift']):
        cat = 'SWIFT / Telex'

    # Reversals / Corrections
    elif any(x in combined for x in ['annulation operation', 'annulation', 'extourne', 'contre-passation',
                                       'afb:   rejet ', 'rejet signature', 'rejet absence']):
        cat = 'Reversal / Correction'

    # Direct Debit / Garnishment
    elif any(x in combined for x in ['prelevement', 'debit direct', 'saisie tra']):
        cat = 'Direct Debit'

    # Domiciliation
    elif any(x in combined for x in ['domiciliation', 'domic', 'fdom ']):
        cat = 'Domiciliation'

    # Collection / Remise
    elif any(x in combined for x in ['remise', 'encaissement']):
        cat = 'Collection / Remise'

    # Card Payment (POS / online — "PMT <merchant> <date>")
    elif any(x in combined for x in ['pmt ', 'an pmt ']):
        cat = 'Card Payment'

    # Loan / Credit
    elif any(x in combined for x in ['pret accorde', 'apl global', 'apl crdit']):
        cat = 'Loan / Credit'

    # Payment Order
    elif any(x in combined for x in ['ordre de paiement', 'payment order']):
        cat = 'Payment Order'

    # Miscellaneous bank entries (catch-all for MISCELLANEOUS SWF codes)
    elif any(x in combined for x in ['miscellaneous']):
        # Try to reclassify based on PY: content
        if any(x in py_lower for x in ['virement', 'vir ']):
            cat = 'Wire Transfer'
        elif any(x in py_lower for x in ['depot especes', 'versement']):
            cat = 'Cash'
        elif any(x in py_lower for x in ['facturation', 'frais', 'commission']):
            cat = 'Fees / Commissions'
        elif 'miscellaneous debit' in combined:
            cat = 'Fees / Commissions'
        else:
            cat = 'Other'

    # Credit divers — general credit entries (collections from third parties)
    elif 'credit divers' in combined:
        cat = 'Collection / Remise'

    # AFB wire received patterns (BOA format: "AFB: 13..." with "recu" or "dol recu")
    elif 'afb:' in combined and ('recu ' in combined or 'dol recu' in combined
                                  or 'vrmt recu' in combined or 'vrt recu' in combined):
        cat = 'Wire Transfer'

    # AFB advances/settlements
    elif 'afb:' in combined and any(x in combined for x in ['avance', 'reliquat', 'reglement',
                                                              'paiement', 'achat ']):
        cat = 'Collection / Remise'

    # AFB inter-agency / solde compte
    elif 'afb:' in combined and ('solde compte' in combined or 'annul.' in combined):
        cat = 'Reversal / Correction'

    # AFB finex fees
    elif 'afb:' in combined and 'fraisfinex' in combined:
        cat = 'Fees / Commissions'

    # AFB circularisation (audit confirmation)
    elif 'afb:' in combined and 'circularisation' in combined:
        cat = 'Fees / Commissions'

    # Unclassified — empty or NONREF with no PY:
    elif (fr.strip() == '' or fr_lower in ['nonref', '']) and not py:
        cat = 'Unclassified'

    # If PY has versement + name pattern (Motif: comments), classify as Cash
    elif 'versement' in py_lower and 'motif:' in comment.lower():
        cat = 'Cash'

    # Regulatory / suspension
    elif any(x in combined for x in ['rgul susp', 'procuration']):
        cat = 'Fees / Commissions'

    else:
        cat = 'Other'

    return cat, description[:80], bnf[:60]

def fmt_date(val):
    """Convert date value to YYYY-MM-DD string."""
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    elif isinstance(val, str):
        try:
            return datetime.strptime(val, '%m/%d/%Y').strftime('%Y-%m-%d')
        except (ValueError, TypeError):
            return str(val)
    return str(val) if val else ''

def classify_flow(t, transfer_indices, idx):
    """Classify transaction into flow type: Collection, Payment, Inter-Account Transfer, Bank Costs, Payroll, Tax."""
    desc = (t.get('description') or '').lower()
    cat = t['category']
    amt = t['amount_xof']

    # --- BANK COSTS ---
    if cat in ['Fees / Commissions', 'Bank Charges', 'Interest / Agios', 'Direct Debit']:
        return 'Bank Costs'
    if cat == 'Tax':
        return 'Tax'

    # --- INTER-ACCOUNT TRANSFERS (keyword-based) ---
    if any(x in desc for x in ['virt inter', 'nivellement', 'appro compte', 'appro caisse',
                                 'transfert interne', 'vir boaweb fav grands', 'vir boaweb recu grands']):
        return 'Inter-Account Transfer'

    # --- INTER-ACCOUNT TRANSFERS (matched pairs) ---
    if idx in transfer_indices:
        if not any(x in desc for x in ['vir.recu:', 'virement recu de la compense',
                                         "virement d'ordre de", 'effet au']):
            return 'Inter-Account Transfer'

    # --- PAYROLL ---
    if cat == 'Payroll / Salary' or any(x in desc for x in ['paiement cnps']):
        return 'Payroll'

    # --- COLLECTIONS (credits from third parties) ---
    if amt > 0:
        return 'Collection'

    # --- PAYMENTS (debits to third parties) ---
    if amt < 0:
        return 'Payment'

    return 'Unclassified'

# Categories excluded from transfer matching (fee/tax noise)
TRANSFER_EXCLUDE_CATS = frozenset({'Fees / Commissions', 'Tax', 'Bank Charges', 'Interest / Agios'})

# Transfer keywords that legitimize same-bank matches
TRANSFER_KEYWORDS = ('virement', 'nivellement', 'appro', 'transfert interne', 'vir boaweb')

def main():
    # Load all files
    all_rows = []
    skipped_rows = 0
    files = [
        ('XOF transactions 1.xlsx', True),
        ('xof transactions.xlsx', False),
        ('XOF 2.xlsx', False),
        ('XOF 3.xlsx', False),
        ('xof 4.xlsx', False),
    ]

    for fname, has_header in files:
        wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if has_header and i == 0:
                continue
            d = parse_row(row)
            if d['Amount'] is None:
                continue
            try:
                amt = float(d['Amount'])
            except (ValueError, TypeError):
                skipped_rows += 1
                continue

            cat, desc, bnf = classify_comment(d['Comment'])
            bank_name = str(d['Bank'] or '')
            country = get_country(bank_name, str(d['AccountNumber'] or ''))

            # Mobile Money: override for MTN Mobile FS CI transactions
            if bank_name == 'MTN Mobile FS CI' and cat in ('Unclassified', 'Other'):
                cat = 'Mobile Money'

            posted = d['PostedDate']
            date_str = fmt_date(posted)
            if isinstance(posted, datetime):
                month_key = posted.strftime('%Y-%m')
            elif isinstance(posted, str):
                try:
                    month_key = datetime.strptime(posted, '%m/%d/%Y').strftime('%Y-%m')
                except (ValueError, TypeError):
                    month_key = 'Unknown'
            else:
                month_key = 'Unknown'

            value_date_str = fmt_date(d['ValueDate'])

            all_rows.append({
                'date': date_str,
                'value_date': value_date_str,
                'month': month_key,
                'account': str(d['Account'] or ''),
                'bank': str(d['Bank'] or ''),
                'country': country,
                'account_type': str(d['AccountType'] or ''),
                'ref': str(d['TransRef'] or ''),
                'amount_xof': amt,
                'amount_usd': round(amt / XOF_PER_USD, 2),
                'category': cat,
                'description': desc,
                'beneficiary': bnf,
                'source': fname,
            })
        wb.close()
        print(f'Loaded {fname}')

    if skipped_rows:
        print(f'WARNING: {skipped_rows} rows skipped due to unparseable amount/date values')

    print(f'\nTotal transactions: {len(all_rows):,}')

    # --- Duplicate detection across source files ---
    GENERIC_REFS = frozenset({'', 'NONREF', 'N/A', 'NA', 'NOREF'})
    seen = set()
    dup_count = 0
    deduped_rows = []
    for r in all_rows:
        ref_val = r['ref'] if r['ref'].upper() not in GENERIC_REFS else r.get('description', '')
        key = (r['date'], r['account'], ref_val, r['amount_xof'])
        if key in seen:
            dup_count += 1
            continue
        seen.add(key)
        deduped_rows.append(r)

    if dup_count:
        print(f'WARNING: {dup_count} duplicate rows detected and removed (same date/account/ref/amount)')
        all_rows = deduped_rows

    # --- Flow classification: find inter-account transfer pairs ---
    by_date_amount = defaultdict(list)
    for i, r in enumerate(all_rows):
        key = (r['date'], abs(r['amount_xof']))
        by_date_amount[key].append(i)

    transfer_indices = set()
    transfer_pairs = []  # List of (debit_idx, credit_idx) for Transfers tab
    for key, indices in by_date_amount.items():
        if len(indices) < 2:
            continue
        debits = [(i, all_rows[i]) for i in indices if all_rows[i]['amount_xof'] < 0
                  and all_rows[i]['category'] not in TRANSFER_EXCLUDE_CATS]
        credits = [(i, all_rows[i]) for i in indices if all_rows[i]['amount_xof'] > 0
                   and all_rows[i]['category'] not in TRANSFER_EXCLUDE_CATS]
        if not debits or not credits:
            continue
        used_credits = set()
        for di, dt in debits:
            for ci, ct in credits:
                if ci in used_credits:
                    continue
                # Must be different accounts AND amounts must match (debit = -credit)
                if dt['account'] != ct['account'] and abs(abs(dt['amount_xof']) - abs(ct['amount_xof'])) < 1:
                    # Same-bank matches require transfer keywords in description
                    if dt['bank'] == ct['bank']:
                        combined_desc = ((dt.get('description') or '') + ' ' + (ct.get('description') or '')).lower()
                        if not any(kw in combined_desc for kw in TRANSFER_KEYWORDS):
                            continue
                    transfer_indices.add(di)
                    transfer_indices.add(ci)
                    transfer_pairs.append((di, ci))
                    used_credits.add(ci)
                    break

    # Apply flow classification to each transaction
    for i, r in enumerate(all_rows):
        r['flow_type'] = classify_flow(r, transfer_indices, i)

    # Aggregate stats
    cat_stats = defaultdict(lambda: {'count': 0, 'debit_xof': 0, 'credit_xof': 0})
    bank_stats = defaultdict(lambda: {'count': 0, 'debit_xof': 0, 'credit_xof': 0})
    month_stats = defaultdict(lambda: {'count': 0, 'debit_xof': 0, 'credit_xof': 0})
    country_stats = defaultdict(lambda: {'count': 0, 'debit_xof': 0, 'credit_xof': 0, 'banks': set()})
    flow_stats = defaultdict(lambda: {'count': 0, 'debit_xof': 0, 'credit_xof': 0})

    for r in all_rows:
        for stats, key in [(cat_stats, r['category']), (bank_stats, r['bank']), (month_stats, r['month']), (flow_stats, r['flow_type'])]:
            stats[key]['count'] += 1
            if r['amount_xof'] < 0:
                stats[key]['debit_xof'] += r['amount_xof']
            else:
                stats[key]['credit_xof'] += r['amount_xof']
        # Country stats
        country_stats[r['country']]['count'] += 1
        country_stats[r['country']]['banks'].add(r['bank'])
        if r['amount_xof'] < 0:
            country_stats[r['country']]['debit_xof'] += r['amount_xof']
        else:
            country_stats[r['country']]['credit_xof'] += r['amount_xof']

    # Build transfer pairs data for Transfers tab (BEFORE sorting all_rows, since indices are pre-sort)
    transfers_list = []
    for di, ci in transfer_pairs:
        dt = all_rows[di]
        ct = all_rows[ci]
        transfers_list.append({
            'date': dt['date'],
            'source_bank': dt['bank'],
            'source_account': dt['account'],
            'source_country': dt['country'],
            'dest_bank': ct['bank'],
            'dest_account': ct['account'],
            'dest_country': ct['country'],
            'amount_xof': abs(dt['amount_xof']),
            'amount_usd': round(abs(dt['amount_xof']) / XOF_PER_USD, 2),
            'description': dt['description'] or ct['description'],
            'debit_ref': dt['ref'],
            'credit_ref': ct['ref'],
        })
    transfers_list.sort(key=lambda x: x['date'])

    # Now safe to sort all_rows
    all_rows.sort(key=lambda x: x['date'])

    print(f'Transfer pairs found: {len(transfers_list):,}')

    # Convert country_stats sets to lists for JSON serialization
    country_stats_json = {}
    for k, v in sorted(country_stats.items(), key=lambda x: -x[1]['count']):
        country_stats_json[k] = {
            'count': v['count'],
            'debit_xof': v['debit_xof'],
            'credit_xof': v['credit_xof'],
            'banks': sorted(list(v['banks'])),
        }

    output = {
        'rate': XOF_PER_USD,
        'total_count': len(all_rows),
        'category_stats': {k: v for k, v in sorted(cat_stats.items(), key=lambda x: -x[1]['count'])},
        'bank_stats': {k: v for k, v in sorted(bank_stats.items(), key=lambda x: -x[1]['count'])},
        'month_stats': {k: v for k, v in sorted(month_stats.items())},
        'country_stats': country_stats_json,
        'flow_stats': {k: v for k, v in sorted(flow_stats.items(), key=lambda x: -x[1]['count'])},
        'transfers': transfers_list,
        'transactions': all_rows,
    }

    with open('ctm_data.json', 'w') as f:
        json.dump(output, f)

    print('Data exported to ctm_data.json')
    print(f'\nCategory breakdown:')
    for cat, s in sorted(cat_stats.items(), key=lambda x: -x[1]['count']):
        print(f"  {cat}: {s['count']:,} txns | Debits: {s['debit_xof']/1e6:,.1f}M XOF | Credits: {s['credit_xof']/1e6:,.1f}M XOF")

    print(f'\nBank breakdown:')
    for bank, s in sorted(bank_stats.items(), key=lambda x: -x[1]['count']):
        print(f"  {bank}: {s['count']:,} txns")

    print(f'\nFlow type breakdown:')
    for flow, s in sorted(flow_stats.items(), key=lambda x: -x[1]['count']):
        d = s['debit_xof']/1e6
        c = s['credit_xof']/1e6
        n = (s['debit_xof'] + s['credit_xof'])/1e6
        print(f"  {flow}: {s['count']:,} txns | Debits: {d:,.1f}M XOF | Credits: {c:,.1f}M XOF | Net: {n:,.1f}M XOF")

if __name__ == '__main__':
    main()
